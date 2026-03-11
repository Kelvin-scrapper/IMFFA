import pdfplumber
import pandas as pd
import os
import re
import json
from datetime import datetime
from openpyxl import Workbook

def clean_value(value_str):
    """Clean extracted values - remove commas, convert -- to None"""
    if not value_str or value_str in ['--', '...', '']:
        return None

    cleaned = str(value_str).replace(',', '').strip()

    # Handle footnote markers like "3/"
    if re.match(r'^\d+/$', cleaned):
        return None

    try:
        return float(cleaned)
    except:
        return None

def date_to_year_week(date_str):
    """Convert date string to year-week format"""
    try:
        for fmt in ['%B %d, %Y', '%b %d, %Y', '%Y-%m-%d', '%m/%d/%Y']:
            try:
                date_obj = datetime.strptime(date_str, fmt)
                year, week, _ = date_obj.isocalendar()
                return f"{year}-{week:02d}"
            except:
                continue
        return None
    except:
        return None

def extract_gra_from_pdf(pdf_path):
    """
    Extract ALL GRA data: aggregates, countries under each arrangement type
    """
    print("\nExtracting GRA data from PDF...")

    all_data = {}
    current_arrangement = None

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()

            if "Current Financial Arrangements (GRA)" not in text:
                continue

            lines = text.split('\n')
            data_started = False

            for line_idx, line in enumerate(lines):
                if not line.strip():
                    continue

                if "Current Financial Arrangements (GRA)" in line:
                    data_started = True
                    continue

                if not data_started:
                    continue

                # Skip header lines
                if any(kw in line for kw in ['Table', 'as of', 'millions of SDRs', 'Member', 'Commited', 'Undrawn', 'Credit Outstanding', 'Period of', 'Effective', 'Expiration']):
                    continue

                # Detect arrangement type headers
                line_upper = line.upper()

                # Standby Arrangement
                if 'STANDBY ARRANGEMENT' in line_upper and '(' in line:
                    match = re.search(r'\((\d+)\)\s+([\d,.-]+|--)\s+([\d,.-]+|--)\s+([\d,.-]+|--)\s+([\d,.-]+|--)', line)
                    if match:
                        all_data['STAND'] = {
                            'AMCOM': clean_value(match.group(2)),
                            'AMUNDRAW': clean_value(match.group(3)),
                            'AMDRAW': clean_value(match.group(4)),
                            'CREDOUTAM': clean_value(match.group(5)),
                            'CREDOUTQUOT': None
                        }
                        current_arrangement = 'STANDBY'
                        print(f"  Found aggregate: STAND")
                        continue

                # Extended Fund Facility
                elif 'EXTENDED FUND FACILITY' in line_upper and '(' in line:
                    match = re.search(r'\((\d+)\)\s+([\d,.-]+|--)\s+([\d,.-]+|--)\s+([\d,.-]+|--)\s+([\d,.-]+|--)', line)
                    if match:
                        all_data['EXTFUNDFAC'] = {
                            'AMCOM': clean_value(match.group(2)),
                            'AMUNDRAW': clean_value(match.group(3)),
                            'AMDRAW': clean_value(match.group(4)),
                            'CREDOUTAM': clean_value(match.group(5)),
                            'CREDOUTQUOT': None
                        }
                        current_arrangement = 'EXTENDED'
                        print(f"  Found aggregate: EXTFUNDFAC")
                        continue

                # Flexible Credit Line
                elif 'FLEXIBLE CREDIT LINE' in line_upper and '(' in line:
                    match = re.search(r'\((\d+)\)\s+([\d,.-]+|--)\s+([\d,.-]+|--)\s+([\d,.-]+|--)\s+([\d,.-]+|--)', line)
                    if match:
                        all_data['FLEXCREDLINE'] = {
                            'AMCOM': clean_value(match.group(2)),
                            'AMUNDRAW': clean_value(match.group(3)),
                            'AMDRAW': clean_value(match.group(4)),
                            'CREDOUTAM': clean_value(match.group(5)),
                            'CREDOUTQUOT': None
                        }
                        current_arrangement = 'FLEXIBLE'
                        print(f"  Found aggregate: FLEXCREDLINE")
                        continue

                # Precautionary Liquidity Line
                elif ('PRECAUTIONARY' in line_upper and 'LIQUIDITY' in line_upper) and '(' in line:
                    match = re.search(r'\((\d+)\)\s+([\d,.-]+|--)\s+([\d,.-]+|--)\s+([\d,.-]+|--)\s+([\d,.-]+|--)', line)
                    if match:
                        all_data['PRELIQLINE'] = {
                            'AMCOM': clean_value(match.group(2)),
                            'AMUNDRAW': clean_value(match.group(3)),
                            'AMDRAW': clean_value(match.group(4)),
                            'CREDOUTAM': clean_value(match.group(5)),
                            'CREDOUTQUOT': None
                        }
                        current_arrangement = 'PRECAUTIONARY'
                        print(f"  Found aggregate: PRELIQLINE")
                        continue

                # Extract country data
                # Pattern: Country name, then numbers, possibly with dates at end
                # Remove arrangement type suffix if present
                line_clean = re.sub(r'(Standby Arrangement|Extended Fund Facility|Flexible Credit Line|Precautionary and Liquidity Line)$', '', line).strip()

                parts = line_clean.split()

                if len(parts) >= 5:
                    has_numbers = any(char.isdigit() for char in line)

                    if has_numbers:
                        country = ""
                        numbers = []

                        for i, part in enumerate(parts):
                            # Check if this is a number or date
                            if re.search(r'^\d', part) or part in ['--', '...']:
                                numbers = parts[i:]
                                country = ' '.join(parts[:i])
                                break

                        if country and len(numbers) >= 4:
                            # Clean country name
                            country_clean = country.strip().upper()
                            # Remove footnotes
                            country_clean = re.sub(r'\s*\d+/$', '', country_clean).strip()

                            # Map to codes
                            country_map = {
                                'ARGENTINA': 'ARG', 'BANGLADESH': 'BGD', 'BENIN': 'BEN',
                                "COTE D'IVOIRE": 'CIV', 'COTE DIVOIRE': 'CIV', 'ECUADOR': 'ECU',
                                'EGYPT': 'EGY', 'EL SALVADOR': 'SLV', 'HONDURAS': 'HND',
                                'JORDAN': 'JOR', 'MAURITANIA': 'MRT', 'MOLDOVA': 'MDA',
                                'PAKISTAN': 'PAK', 'PAPUA NEW GUINEA': 'PNG', 'SENEGAL': 'SEN',
                                'SEYCHELLES': 'SYC', 'SRI LANKA': 'LKA', 'UKRAINE': 'UKR',
                                'CHILE': 'CHL', 'COLOMBIA': 'COL', 'COSTA RICA': 'CRI',
                                'MEXICO': 'MEX', 'MOROCCO': 'MAR', 'ARMENIA': 'ARM',
                                'GEORGIA': 'GEO', 'KOSOVO': 'XKX', 'SERBIA': 'SRB',
                                'BARBADOS': 'BRB', 'CAMEROON': 'CMR', 'GABON': 'GAB',
                                'SURINAME': 'SUR', 'NORTH MACEDONIA': 'MKD'
                            }

                            country_code = country_map.get(country_clean)

                            if country_code:
                                # Handle "3/" footnote marker - skip it when parsing
                                num_start = 0
                                for idx, num in enumerate(numbers):
                                    if re.match(r'^\d+/$', num):  # This is "3/" or similar
                                        num_start = idx + 1
                                        break

                                # Adjust numbers array if we skipped footnote
                                actual_numbers = numbers[num_start:] if num_start > 0 else numbers

                                all_data[country_code] = {
                                    'AMCOM': clean_value(actual_numbers[0]) if len(actual_numbers) > 0 else None,
                                    'AMUNDRAW': clean_value(actual_numbers[1]) if len(actual_numbers) > 1 else None,
                                    'AMDRAW': clean_value(actual_numbers[2]) if len(actual_numbers) > 2 else None,
                                    'CREDOUTAM': clean_value(actual_numbers[3]) if len(actual_numbers) > 3 else None,
                                    'CREDOUTQUOT': clean_value(actual_numbers[4]) if len(actual_numbers) > 4 else None
                                }
                                print(f"  Found: {country_clean} ({country_code})")

                # Check for "Total Current GRA Arrangements" line
                if 'TOTAL CURRENT GRA' in line_upper:
                    match = re.search(r'\((\d+)\)\s+([\d,.-]+|--)\s+([\d,.-]+|--)\s+([\d,.-]+|--)\s+([\d,.-]+|--)', line)
                    if match:
                        all_data['TOTAL'] = {
                            'AMCOM': clean_value(match.group(2)),
                            'AMUNDRAW': clean_value(match.group(3)),
                            'AMDRAW': clean_value(match.group(4)),
                            'CREDOUTAM': clean_value(match.group(5)),
                            'CREDOUTQUOT': None
                        }
                        print(f"  Found: Total Current GRA Arrangements")
                        continue

    return all_data

def extract_fcc_from_pdf(pdf_path):
    """Extract FCC data"""
    print("\nExtracting FCC data from PDF...")

    fcc_data = {}

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()

            if "Forward Commitment Capacity (FCC)" not in text or "SDR USD" not in text:
                continue

            lines = text.split('\n')
            in_section = False

            for line in lines:
                if not line.strip():
                    continue

                if "Forward Commitment Capacity (FCC)" in line and "SDR USD" in line:
                    in_section = True
                    continue

                if in_section and ("Rates" in line or "1/ Figures" in line or "1 SDR" in line):
                    break

                if not in_section:
                    continue

                line_clean = ' '.join(line.split())
                # Match with optional footnote markers like "5/"
                match = re.search(r'^(.+?)\s+([\d,.-]+|--)\s+([\d,.-]+|--)$', line_clean)

                if match:
                    metric_text = match.group(1).strip()
                    sdr_value = clean_value(match.group(2))
                    usd_value = clean_value(match.group(3))

                    # Clean metric name - remove footnotes, roman numerals, letters
                    metric_clean = re.sub(r'^[IVX]+\.\s*', '', metric_text)
                    metric_clean = re.sub(r'^\([a-z]\)\s*', '', metric_clean)
                    metric_clean = re.sub(r'\s+\d+/$', '', metric_clean)  # Remove footnotes like "5/"
                    metric_clean = metric_clean.strip()

                    # Match to codes - ORDER MATTERS! More specific terms first
                    metric_map = {
                        'Usable resources (a) + (b)': 'USRES',
                        'Fund quota resources': 'FUNDQOUTRES',
                        'Fund borrowed resources': 'FUNDBORRES',
                        'Undrawn balances under GRA lending commitments': 'UNDRAWBALCOM',
                        'Non-precautionary': 'NONPRECAUT',  # Must be before 'Precautionary'
                        'Precautionary': 'PRECAUT',
                        'Uncommitted usable resources ( I - II )': 'UNCOMUSRES',
                        'Repurchases one-year forward': 'REPURCHONEYFOR',
                        'Repayments of borrowed resources one-year forward': 'REPAYONEYFOR',
                        'Prudential balance': 'PRUDBAL',
                        'Forward commitment capacity ( III + IV - V - VI )': 'FORCOMCAP',
                        'From Quota resources': 'QUOTRES',
                        'From NAB resources': 'NABRES',
                        'From Bilateral Borrowed resources': 'BILBORRES'
                    }

                    for key, code in metric_map.items():
                        # Match exactly or by normalized version
                        key_lower = key.lower()
                        metric_lower = metric_clean.lower()

                        # Exact match first
                        if key_lower == metric_lower:
                            if code not in fcc_data:
                                fcc_data[code] = {'SDR': sdr_value, 'USD': usd_value}
                                print(f"  Found: {metric_clean[:50]} -> {code}")
                            break

                        # For "Uncommitted usable resources", match without roman numerals
                        if '( i - ii )' in key_lower and 'uncommitted usable resources' in metric_lower:
                            if code not in fcc_data:
                                fcc_data[code] = {'SDR': sdr_value, 'USD': usd_value}
                                print(f"  Found: {metric_clean[:50]} -> {code}")
                            break

            break

    return fcc_data

def extract_memitem_from_pdf(pdf_path):
    """Extract Memo Items (Outstanding Disbursements and Total)"""
    print("\nExtracting Memo Items from PDF...")

    memitem_data = {}

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages[:5]:
            text = page.extract_text()

            if "Memorandum Items" not in text:
                continue

            lines = text.split('\n')

            for i, line in enumerate(lines):
                # Look for the specific memo item lines
                # "Credit outstanding from members without current arrangement and outright disbursements. 8,182"
                if "Credit outstanding from members without current arrangement" in line:
                    # Extract the number at the end
                    match = re.search(r'(\d{1,3},?\d{3,})\.?\s*$', line)
                    if match:
                        val = clean_value(match.group(1))
                        memitem_data['OUTDIS'] = val
                        print(f"  Found OUTDIS: {val}")

                # "Total Credit Outstanding. 91,000"
                elif "Total Credit Outstanding" in line and "." in line:
                    match = re.search(r'(\d{1,3},?\d{3,})\.?\s*$', line)
                    if match:
                        val = clean_value(match.group(1))
                        memitem_data['TOTAL'] = val
                        print(f"  Found TOTAL: {val}")

    return memitem_data

def build_output(year_week, gra_data, fcc_data, memitem_data, reference_columns):
    """Build output matching reference structure"""
    print("\nBuilding output DataFrame...")

    data_dict = {}

    for col_tuple in reference_columns:
        col_name = col_tuple[0]

        if 'Unnamed' in col_name:
            data_dict[col_tuple] = year_week
            continue

        parts = col_name.split('.')
        value = None

        if len(parts) >= 4:
            table = parts[1]
            metric = parts[2]
            entity = parts[3]

            if table == 'CURFIN':
                if metric == 'MEMITEM' and entity in memitem_data:
                    value = memitem_data[entity]
                elif entity in gra_data and metric in gra_data[entity]:
                    value = gra_data[entity][metric]

            elif table == 'FORCOMCAP' and metric in fcc_data:
                if entity in fcc_data[metric]:
                    value = fcc_data[metric][entity]

        data_dict[col_tuple] = value

    data_values = list(data_dict.values())
    df = pd.DataFrame([data_values], columns=reference_columns)

    print(f"  Total columns: {len(df.columns)}")
    print(f"  Non-null values: {df.notna().sum().sum()}")

    return df

def process_single_pdf(pdf_path, reference_columns):
    """Process a single PDF and return data row"""
    print(f"\n  Processing: {os.path.basename(pdf_path)}")

    # Extract date
    year_week = None
    with pdfplumber.open(pdf_path) as pdf:
        text = pdf.pages[0].extract_text()
        match = re.search(r'as of\s+([A-Za-z]+\s+\d{1,2},\s+\d{4})', text)
        if match:
            report_date = match.group(1)
            year_week = date_to_year_week(report_date)
            print(f"    Date: {report_date} -> {year_week}")

    # Extract data
    gra_data = extract_gra_from_pdf(pdf_path)
    fcc_data = extract_fcc_from_pdf(pdf_path)
    memitem_data = extract_memitem_from_pdf(pdf_path)

    # Build data row
    data_dict = {}
    for col_tuple in reference_columns:
        col_name = col_tuple[0]

        if 'Unnamed' in col_name:
            data_dict[col_tuple] = year_week
            continue

        parts = col_name.split('.')
        value = None

        if len(parts) >= 4:
            table = parts[1]
            metric = parts[2]
            entity = parts[3]

            if table == 'CURFIN':
                if metric == 'MEMITEM' and entity in memitem_data:
                    value = memitem_data[entity]
                elif entity in gra_data and metric in gra_data[entity]:
                    value = gra_data[entity][metric]

            elif table == 'FORCOMCAP' and metric in fcc_data:
                if entity in fcc_data[metric]:
                    value = fcc_data[metric][entity]

        data_dict[col_tuple] = value

    data_values = list(data_dict.values())
    non_null = sum([1 for v in data_values if pd.notna(v)])
    print(f"    Extracted: {non_null} non-null values")

    return data_values

def create_template_from_config():
    """Create reference template from config.json if needed"""
    print("\n[INFO] Creating reference template from config.json...")

    with open('config.json', 'r') as f:
        config = json.load(f)

    sorted_mappings = sorted(config['column_mapping'].items(), key=lambda x: x[1]['order'])

    columns = []
    for col_name, col_info in sorted_mappings:
        columns.append((col_name, col_info['display_name'].replace('_', ' ')))

    date_col = [(config['date_format']['column_name'], f"Date ({config['date_format']['type']})")]
    all_columns = date_col + columns

    output_file = "IMFFA_DATA_.xlsx"

    # Create Excel file with two header rows manually
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active

    # Write first header row (technical names)
    for idx, (tech_name, _) in enumerate(all_columns, start=1):
        ws.cell(row=1, column=idx, value=tech_name)

    # Write second header row (display names)
    for idx, (_, display_name) in enumerate(all_columns, start=1):
        ws.cell(row=2, column=idx, value=display_name)

    # Write example data row
    ws.cell(row=3, column=1, value=config['date_format']['example'])

    wb.save(output_file)
    print(f"  [SUCCESS] Template created: {output_file} ({len(all_columns)} columns)")
    return output_file

def main():
    print("="*80)
    print("IMF FINANCIAL ACTIVITIES DATA EXTRACTOR - BATCH MODE")
    print("="*80)

    # Load or create reference
    excel_file = "IMFFA_DATA_.xlsx"

    if not os.path.exists(excel_file):
        if os.path.exists('config.json'):
            print("\n[INFO] Creating reference template from config.json...")
            excel_file = create_template_from_config()
        else:
            print("\n[ERROR] Neither IMFFA_DATA_.xlsx nor config.json found!")
            return

    reference_columns = pd.read_excel(excel_file, sheet_name=0, header=[0, 1]).columns
    print(f"\nReference loaded: {len(reference_columns)} columns")

    # Find all PDFs
    pdf_files = []
    for folder in ["Downloads", "output"]:
        if os.path.exists(folder):
            for file in os.listdir(folder):
                if file.endswith(".pdf"):
                    pdf_files.append(os.path.join(folder, file))

    if not pdf_files:
        print("\nERROR: No PDF files found in Downloads/ or output/")
        return

    # Sort by modification time (newest first)
    pdf_files.sort(key=os.path.getmtime, reverse=True)

    print(f"\nFound {len(pdf_files)} PDF file(s) to process:")
    for i, pdf in enumerate(pdf_files, 1):
        print(f"  {i}. {os.path.basename(pdf)}")

    # Process all PDFs
    print(f"\n{'='*80}")
    print("EXTRACTING DATA FROM ALL PDFS...")
    print(f"{'='*80}")

    all_data_rows = []
    for pdf_path in pdf_files:
        try:
            data_row = process_single_pdf(pdf_path, reference_columns)
            all_data_rows.append(data_row)
        except Exception as e:
            print(f"    ERROR: {str(e)}")
            continue

    if not all_data_rows:
        print("\nNo data extracted")
        return

    # Save combined output
    print(f"\n{'='*80}")
    print(f"CREATING COMBINED OUTPUT")
    print(f"{'='*80}")

    # Create output folder
    os.makedirs('output', exist_ok=True)
    output_file = os.path.join('output', 'IMFFA_DATA_OUTPUT.xlsx')

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Write headers
    for col_idx, col_tuple in enumerate(reference_columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_tuple[0])
        ws.cell(row=2, column=col_idx, value=col_tuple[1])

    # Write all data rows
    for row_idx, data_row in enumerate(all_data_rows, start=3):
        for col_idx, value in enumerate(data_row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(output_file)

    print(f"\n{'='*80}")
    print(f"SUCCESS! Batch output saved to: {output_file}")
    print(f"  Total PDFs processed: {len(all_data_rows)}")
    print(f"  Total columns: {len(reference_columns)}")
    print(f"  Output rows: {len(all_data_rows)} (plus 2 header rows)")
    print(f"{'='*80}")

if __name__ == "__main__":
    main()